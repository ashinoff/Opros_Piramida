# -*- coding: utf-8 -*-
"""Потоковый парсер выгрузки «Опрос ПУ» из Пирамиды.

Формат стабилен: лист «ПОДРОБНО», шапка до строки 10 (заголовки в строке 10),
данные с 11-й строки. Колонки:
B ПУ | C Тип ПУ | D Серийный | E Фазность | F Маршрут | G Дата опроса |
H Точка учёта | I Область | J РЭС | K Фидер | L ТП | M Тип абонента |
N Тип ТУ | O Отключен | P Собираемость | Q Угасание | R Модуляция | S Тип данных
"""
import re
import openpyxl

HEADER_ROW_MARKER = ("ПУ", "Тип ПУ")  # значения B и C в строке заголовков

# --- производители: по началу наименования типа ПУ (3–7 знаков) ---
VENDOR_RULES = [
    ("РиМ", "РиМ"),
    ("Матрица", "Матрица"),
    ("Меркурий", "Меркурий"),
    ("Нартис", "Нартис"),
    ("СЕ", "Энергомера (СЕ)"),
    ("ЦЭ", "Энергомера (ЦЭ)"),
    ("AD", "ADDAX (AD)"),
    ("Миртек", "Миртек"),
    ("МИРТЕК", "Миртек"),
    ("МИР ", "МИР"),
    ("Фобос", "Фобос"),
    ("ФОБОС", "Фобос"),
    ("ПСЧ", "Нижегородец (ПСЧ)"),
    ("СЭБ", "Нижегородец (СЭБ)"),
    ("Квант", "Квант"),
    ("КВАНТ", "Квант"),
    ("Милур", "Милур"),
    ("NP", "Матрица"),
    ("Каскад", "Каскад"),
    ("Вектор", "Вектор"),
    ("Пульсар", "Пульсар"),
    ("Энергомера", "Энергомера"),
]


def detect_vendor(type_name: str) -> str:
    t = (type_name or "").strip()
    for prefix, vendor in VENDOR_RULES:
        if t.startswith(prefix):
            return vendor
    # запасной вариант — первое слово
    return t.split(" ")[0] if t else "Не определён"


def is_spodes(type_name: str) -> bool:
    return "СПОДЭС" in (type_name or "").upper()


# --- классификация маршрута опроса ---
def classify_route(route: str):
    """Возвращает (класс, идентификатор ведущего устройства)."""
    r = (route or "").strip()
    if not r:
        return "Нет маршрута", ""
    low = r.lower()
    if low.startswith("rootrouter"):
        cls = "RootRouter"
    elif low.startswith("rtr"):
        cls = "RTR"
    elif low.startswith("мкс"):
        cls = "МКС"
    elif low.startswith("без порта"):
        return "Без порта", ""
    elif "gsm" in low or "мтс" in low or "мегафон" in low or "билайн" in low \
            or re.match(r"^(\+?7|8)\d{10}", r):
        cls = "GSM"
    elif re.match(r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}", r):
        cls = "IP-шлюз"
    else:
        cls = "Прочее"
    # идентификатор устройства: часть до " - " (модель + №)
    device = r.split(" - ")[0].strip() if " - " in r else r
    device = device[:120]
    return cls, device


PERIOD_RE = re.compile(r"с\s+([\d.]+)\s+по\s+([\d.]+)")


def parse_file(path: str):
    """Генератор: сначала dict с метаданными {'meta': ...}, затем строки-словари."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # ищем лист с подробными данными
    sheet = None
    for name in wb.sheetnames:
        if "ПОДРОБ" in name.upper():
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb[wb.sheetnames[0]]

    meta = {"period_from": "", "period_to": ""}
    header_found = False
    rows_started = False

    for row in sheet.iter_rows(values_only=True):
        b = row[1] if len(row) > 1 else None
        if not rows_started:
            if isinstance(b, str) and "за период" in b:
                m = PERIOD_RE.search(b)
                if m:
                    meta["period_from"], meta["period_to"] = m.group(1), m.group(2)
            if b == HEADER_ROW_MARKER[0] and (len(row) > 2 and row[2] == HEADER_ROW_MARKER[1]):
                header_found = True
                rows_started = True
                yield {"meta": meta}
                continue
            continue
        if b is None:
            continue
        if len(row) < 19:
            row = tuple(row) + (None,) * (19 - len(row))
        type_name = str(row[2] or "").strip()
        serial = str(row[3] or "").strip()
        if not serial and not type_name:
            continue
        route_raw = str(row[5] or "").strip()
        route_class, route_device = classify_route(route_raw)
        poll_date = row[6]
        yield {
            "type_name": type_name,
            "serial": serial,
            "vendor": detect_vendor(type_name),
            "is_spodes": is_spodes(type_name),
            "phase": str(row[4] or "").strip(),
            "route_raw": route_raw[:200],
            "route_class": route_class,
            "route_device": route_device,
            "poll_date": poll_date.strftime("%d.%m.%Y") if hasattr(poll_date, "strftime") else str(poll_date or ""),
            "tu_path": str(row[7] or "").strip(),
            "res": str(row[9] or "").strip() or "Без РЭС",
            "feeder": str(row[10] or "").strip(),
            "tp": str(row[11] or "").strip(),
            "abonent_type": str(row[12] or "").strip(),
            "tu_type": str(row[13] or "").strip(),
            "disconnected": bool(str(row[14] or "").strip()),
            "collected": str(row[15] or "").strip() == "Собирается",
            "fading": bool(str(row[16] or "").strip()),
            "modulation": str(row[17] or "").strip(),
            "data_type": str(row[18] or "").strip(),
        }
    if not header_found:
        raise ValueError("Не найдена строка заголовков (ПУ / Тип ПУ). "
                         "Проверьте, что это выгрузка «Опрос ПУ» из Пирамиды.")
